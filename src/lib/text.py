import string
import os
import sys
import csv
import json
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


def normalize(text: str, *, casefold: bool = True, yo2e: bool = True) -> str:
    s = text
    if casefold:
        s = s.casefold()
    if yo2e:
        s = s.replace('ё', 'е').replace('Ё', 'Е')
    s = s.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    s = s.strip()
    while '  ' in s:
        s = s.replace('  ', ' ')
    return s

def tokenize(text: str) -> list[str]:
    s = text
    cyrillic_lower_letters = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'
    alp = cyrillic_lower_letters + string.ascii_lowercase + string.digits
    for i in range(len(s)):
        if s[i] in alp:
            continue
        else:
            s = s[:i] + ' ' + s[i+1:]
    while '  ' in s:
        s = s.replace('  ', ' ')
    s = s.split()
    return s

def count_freq(tokens: list[str]) -> dict[str, int]:
    d = {}
    tokens_set = set(tokens)
    for key in tokens_set:
        d[key] = tokens.count(key)
    return d

def top_n(freq: dict[str, int], n: int = 5) -> list[tuple[str, int]]:
    d = freq
    d = sorted(d.items(), key=lambda para: (-para[1], para[0]))
    return d[:n]

def ensure_directory_exists(file_path: str) -> None:
    """Создает директорию для файла, если она не существует"""
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory)

def table_output(text: str, top: int = 5):
    # Получаем список всех слов
    text_corrected = tokenize(normalize(text))
    # Считаем общее кол-во слов
    count_words = len(text_corrected)
    # Получаем словарь уникальных слов
    dict_words = count_freq(text_corrected)
    # Считаем кол-во уникальных слов
    count_words_unique = len(dict_words)
    # Сортируем словарь по кол-ву слов
    dict_words_sort = top_n(dict_words, top)

    print(f'Всего слов: {count_words}')
    print(f'Уникальных слов: {count_words_unique}')
    print()
    print(f'Топ {top}:')
    
    # Находим максимальную длину слова и частоты в топ-N
    max_word_len = max(len(word) for word, _ in dict_words_sort[:top])
    max_count_len = max(len(str(count)) for _, count in dict_words_sort[:top])
    
    # Устанавливаем ширину колонок
    col_width = max(15, max_word_len + 2, max_count_len + 2)
    
    k = 0
    print(f'{"слово:":^{col_width}} |{"частота":^{col_width}}')
    print(f"{'-' * col_width}-|-{'-' * col_width}")
    for word, counts in dict_words_sort:
        if k == top:
            break
        k += 1
        print(f'{word:^{col_width}} |{counts:^{col_width}}')


# Добавляем путь для импорта func1
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

try:
    from lib.func1 import is_file_empty
except ImportError:
    # Если func1 нет, создаем простую версию
    def is_file_empty(file_path: str) -> bool:
        """Проверяет пустой ли файл"""
        return os.path.getsize(file_path) == 0

def json_to_csv(json_path: str, csv_path: str) -> None:
    # Проверка json файла
    json_file = Path(json_path)
    if not json_file.exists():
        raise FileNotFoundError(f'JSON файл не найден: {json_path}')
    
    if json_file.suffix.lower() != '.json':
        raise TypeError(f'Файл должен быть в формате json')
    
    if is_file_empty(json_path):
        raise FileNotFoundError(f'Файл {json_path} пустой')
    
    ensure_directory_exists(csv_path)
    
    # Проверка csv файла
    try:
        csv_file = Path(csv_path)
    except:
        raise FileNotFoundError(f'csv файл не найден: {csv_path}')
    
    if csv_file.suffix.lower() != '.csv':
        raise TypeError(f'Файл должен быть в формате csv')
    
    try:
        with open(json_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)
    except json.JSONDecodeError as e:
        raise ValueError(f'Ошибка парсинга JSON: {e}')
    
    # Проверяем, что json состоит из списков
    if not isinstance(data, list):
        raise ValueError("JSON должен содержать список объектов")
    
    # Проверяем структуры данных
    if not all(isinstance(item, dict) for item in data):
        raise ValueError("Все элементы в JSON должны быть словарями")
    
    fieldnames = set()
    for item in data:
        fieldnames.update(item.keys())
    
    try:
        with open(csv_path, 'w', encoding='utf-8', newline='') as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            writer.writeheader()
            for item in data:
                row = {key: item.get(key, '') for key in fieldnames}
                writer.writerow(row)
    except Exception as e:
        raise ValueError(f"Ошибка записи CSV: {e}")

def csv_to_json(csv_path: str, json_path: str) -> None:
    csv_file = Path(csv_path)
    ensure_directory_exists(json_path)
    json_file = Path(json_path)
    
    try:
        data = []     
        with open(csv_file, 'r', encoding='utf-8') as f:
            # Проверяем заголовки 
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            try:            
                has_header = sniffer.has_header(sample)
            except csv.Error:
                has_header = True           
            if not has_header:
                raise ValueError("CSV файл, вероятно, не содержит заголовок (эвристический подход)")	
            
            # Пытаемся прочитать CSV с заголовками
            csv_reader = csv.DictReader(f)
            
            # Проверяем, есть ли поле fieldnames (заголовки)
            if csv_reader.fieldnames is None:
                raise ValueError("CSV файл не содержит заголовков!")
            # Проверяем, что все заголовки уникальны
            if len(csv_reader.fieldnames) != len(set(csv_reader.fieldnames)):
                raise ValueError("Заголовки CSV содержат дубликаты!")
            
            # Читаем данные
            for row in csv_reader:
                data.append(row)
        
        # Сохраняем JSON
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        
    except Exception as e:
        print(f"Ошибка: {e}")

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    # Добавляем путь csv и xlsx файла
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)
    
    # Проверяем существует ли файл
    if not csv_file.exists():
        raise FileNotFoundError(f'Файл {csv_path} не найден')
    
    # Проверяем пустой файл или нет
    if is_file_empty(csv_file):
        raise ValueError(f'Файл {csv_path} пустой')
    
    # Проверяем формат файлов
    if csv_file.suffix.lower() != '.csv':
        raise TypeError(f'файл {csv_path} не формата csv')

    # Создаем директорию для XLSX файла если её нет
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)

    # Читаем CSV и конвертируем в XLSX
    try:
        # Создаем новую книгу Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"

        with open(csv_file, 'r', encoding='utf-8', newline='') as file:
            csv_reader = csv.reader(file)

            # При помощи циклов записываем данные в xlsx
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Задаем ширину колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column) 
            
            for cell in column:
                if cell.value:  # Проверяем что значение не None
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Устанавливаем ширину колонки (не менее 8 символов)
            column_width = max(max_length + 2, 8)
            worksheet.column_dimensions[column_letter].width = column_width

        # Сохраняем файл
        workbook.save(xlsx_file)
        print(f"Успешно создан XLSX файл: {xlsx_path}")
        
    except Exception as e:
        raise ValueError(f'Не получилось считать CSV файл: {e}')
