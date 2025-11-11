import openpyxl, os, sys, csv
from pathlib import Path
from openpyxl.utils import get_column_letter

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from lib.func1 import is_file_empty

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    # Добавляем путь csv и xlsx файла
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)
    
    # Проверяем существует ли файл
    if not csv_file.exists():
        raise FileNotFoundError(f'Файл {csv_path} не найден')
    
    # Проверяем пустой файл или нет
    if is_file_empty(csv_file):
        raise ValueError(f'Файл {csv_path} пустой')
    
    # Проверяем формат файлов
    if csv_file.suffix.lower() != '.csv':
        raise TypeError(f'файл {csv_path} не формата csv')

    # Создаем директорию для XLSX файла если её нет
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)

    # Читаем CSV и конвертируем в XLSX
    try:
        # Создаем новую книгу Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"

        with open(csv_file, 'r', encoding='utf-8', newline='') as file:
            csv_reader = csv.reader(file)

            # При помощи циклов записываем данные в xlsx
            for row_idx, row in enumerate(csv_reader, 1):
                for col_idx, value in enumerate(row, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Задаем ширину колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column) 
            
            for cell in column:
                if cell.value:  # Проверяем что значение не None
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Устанавливаем ширину колонки (не менее 8 символов)
            column_width = max(max_length + 2, 8)
            worksheet.column_dimensions[column_letter].width = column_width

        # Сохраняем файл
        workbook.save(xlsx_file)
        print(f"Успешно создан XLSX файл: {xlsx_path}")
        
    except Exception as e:
        raise ValueError(f'Не получилось считать CSV файл: {e}')

try:
    csv_to_xlsx("python_labs/data/samples/people.csv", "python_labs/data/out/people.xlsx")
    csv_to_xlsx("python_labs/data/samples/cities.csv", "python_labs/data/out/cities.xlsx")  
except Exception as e:
    print(f"Ошибка: {e}")